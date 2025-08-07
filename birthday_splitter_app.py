import streamlit as st
import pandas as pd
import os
import tempfile
import zipfile
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

st.set_page_config(page_title="Birthday Excel Splitter", layout="centered")
st.title("🎂 Split Excel by Birthday Day")
st.write("Upload your Excel file (with `Dge`, `telefoni`, and `segment` columns) and download files grouped by birthday day.")
uploaded_file = st.file_uploader("Upload Excel file (.xlsx)", type="xlsx")
if uploaded_file:
    with tempfile.TemporaryDirectory() as tmpdir:
        df = pd.read_excel(uploaded_file, sheet_name='Sheet1')

        # Clean
        df = df.dropna(subset=['Dge', 'telefoni', 'segment'])
        df = df[~df['telefoni'].astype(str).str.strip().eq('')]
        df = df[~df['segment'].astype(str).str.strip().eq('')]
        df['Dge'] = df['Dge'].astype(str).str.strip()
        df = df[df['Dge'].str.isdigit()]

        # Output folder
        output_folder = os.path.join(tmpdir, "by_day")
        os.makedirs(output_folder, exist_ok=True)

        for day, group in df.groupby('Dge'):
            filename = f'day_{day.zfill(2)}.xlsx'
            output_path = os.path.join(output_folder, filename)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                group.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                for i, column in enumerate(group.columns, 1):
                    max_len = max(group[column].astype(str).map(len).max(), len(str(column)))
                    worksheet.column_dimensions[get_column_letter(i)].width = max_len + 2

        # Zip the folder
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zipf:
            for filename in os.listdir(output_folder):
                full_path = os.path.join(output_folder, filename)
                zipf.write(full_path, arcname=filename)

        zip_buffer.seek(0)
        st.success("✅ All files grouped and saved!")

        st.download_button(
            label="📦 Download ZIP of Day-wise Excel Files",
            data=zip_buffer,
            file_name="birthday_groups.zip",
            mime="application/zip"
        )
